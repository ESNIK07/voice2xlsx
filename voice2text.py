import os
import speech_recognition as sr
import re
from number_parser import parse_number
from openpyxl import Workbook, load_workbook
from datetime import datetime

# 📂 Ruta donde se guardarán los archivos
CARPETA_REGISTROS = "registros_financieros"

# ✅ Crear carpeta si no existe
if not os.path.exists(CARPETA_REGISTROS):
    os.makedirs(CARPETA_REGISTROS)

# 🔧 Configuración de reconocimiento de voz
recognizer = sr.Recognizer()
recognizer.pause_threshold = 1.5
recognizer.energy_threshold = 300
recognizer.dynamic_energy_threshold = True

# Diccionario para corregir errores comunes
correcciones = {
    "bendy": "vendí",
    "bendice": "vendí",
    "compre": "compré",
    "compres": "compré",
    "vendi": "vendí",
    "vende": "vendí",
    "ofresi": "ofrecí"
}

# 🛠️ Corregir errores comunes de transcripción
def corregir_errores(texto):
    palabras = texto.split()
    corregido = [correcciones.get(palabra, palabra) for palabra in palabras]
    return " ".join(corregido)

# 🎙️ Escuchar comando de voz
def escuchar_comando():
    with sr.Microphone() as source:
        print("🎤 Escuchando... di tu comando.")
        audio = recognizer.listen(source)

        try:
            texto = recognizer.recognize_google(audio, language="es-CO")
            print(f"🔊 Entendí: {texto}")
            texto_corregido = corregir_errores(texto.lower())
            print(f"🛠️ Texto corregido: {texto_corregido}")
            return texto_corregido
        except sr.UnknownValueError:
            print("❌ No se pudo entender el audio.")
        except sr.RequestError as e:
            print(f"⚠️ Error de conexión: {e}")

# 🔢 Convertir números escritos a dígitos
def convertir_numeros(texto):
    palabras = texto.split()
    convertido = []

    for palabra in palabras:
        try:
            numero = parse_number(palabra)
            if numero is not None:
                convertido.append(str(numero))
            else:
                convertido.append(palabra)
        except:
            convertido.append(palabra)
    
    return " ".join(convertido)

# 💰 Limpiar el formato del valor (eliminando separadores de miles y verificando formatos)
def limpiar_valor(valor_str):
    # Eliminar símbolos de moneda y espacios
    valor_str = valor_str.replace("$", "").replace(" ", "")

    # Caso 1: Si el número contiene comas y puntos (asumimos coma como decimal)
    if "," in valor_str and "." in valor_str:
        valor_str = valor_str.replace(".", "")  # Eliminar puntos (separadores de miles)
        valor_str = valor_str.replace(",", ".")  # Convertir coma decimal a punto decimal
        valor = float(valor_str)
    # Caso 2: Solo comas (asumimos separadores de miles)
    elif "," in valor_str:
        valor_str = valor_str.replace(",", "")
        valor = int(valor_str)
    # Caso 3: Solo puntos (asumimos separadores de miles)
    elif "." in valor_str:
        valor_str = valor_str.replace(".", "")
        valor = int(valor_str)
    # Caso 4: Número limpio
    else:
        valor = int(valor_str)

    return int(valor)  # Devolver siempre un número entero

# ✅ Confirmar antes de guardar la transacción
def confirmar_transaccion(operacion, valor):
    print(f"\n📝 Operación detectada: {operacion.upper()}")
    print(f"💰 Valor detectado: {valor} COP")
    confirmar = input("✅ ¿Deseas guardar esta transacción? (s/n): ").strip().lower()
    return confirmar == 's'

# 🔍 Identificar si es compra o venta
def identificar_operacion(accion):
    acciones_compra = ["compré", "compramos", "se compró", "adquirí", "adquirimos", "obtuvimos", "obtuve"]
    acciones_venta = ["vendí", "vendimos", "se vendió", "ofrecí", "ofrecimos", "despaché", "despachamos"]

    if accion in acciones_compra:
        return "compra"
    elif accion in acciones_venta:
        return "venta"
    else:
        return None

# 📊 Guardar la transacción en Excel
def guardar_transaccion(operacion, valor):
    ahora = datetime.now()
    fecha_actual = ahora.strftime("%d-%m-%Y")
    hora_actual = ahora.strftime("%H:%M:%S")
    mes_actual = ahora.strftime("%B_%Y").lower()

    nombre_archivo = f"{CARPETA_REGISTROS}/finanzas_{mes_actual}.xlsx"

    # 📂 Abrir o crear archivo mensual
    if os.path.exists(nombre_archivo):
        libro = load_workbook(nombre_archivo)
    else:
        libro = Workbook()

    # 📄 Crear o seleccionar la hoja del día
    if fecha_actual in libro.sheetnames:
        hoja = libro[fecha_actual]
    else:
        hoja = libro.create_sheet(title=fecha_actual)
        # 🏷️ Crear estructura de la hoja
        hoja.append(["Compras", "", "Ventas", ""])
        hoja.append(["Fecha y hora", "Valor (COP)", "Fecha y hora", "Valor (COP)"])

    # 📝 Registrar la transacción en la sección correcta
    if operacion == "compra":
        columna_fecha = 'A'
        columna_valor = 'B'
    else:
        columna_fecha = 'C'
        columna_valor = 'D'

    fila = 3  # Comenzar después de los encabezados
    while hoja[f"{columna_fecha}{fila}"].value:
        fila += 1

    hoja[f"{columna_fecha}{fila}"] = f"{fecha_actual} {hora_actual}"
    hoja[f"{columna_valor}{fila}"] = valor

    # ➕ Agregar totales
    hoja[f"A{fila + 2}"] = "Total Compras:"
    hoja[f"B{fila + 2}"] = f"=SUM(B3:B{fila})"
    hoja[f"C{fila + 2}"] = "Total Ventas:"
    hoja[f"D{fila + 2}"] = f"=SUM(D3:D{fila})"

    libro.save(nombre_archivo)
    print(f"✅ Transacción registrada en {nombre_archivo} - Hoja: {fecha_actual}")

# 🔍 Interpretar comando de voz
def interpretar_comando(texto):
    texto_convertido = convertir_numeros(texto)
    print(f"🔢 Texto convertido: {texto_convertido}")

    patron = r"(compr[ée]?|compramos|se compr[oó]|adquir[ií]|adquirimos|obtuve|obtuvimos|vend[íi]|vendimos|se vendi[oó]|ofrec[íi]|ofrecimos|despach[ée]|despachamos) (?:por el valor de|por)? ?\$?([\d.,]+)"
    resultado = re.search(patron, texto_convertido.lower())

    if not resultado:
        print("❌ No se pudo interpretar el comando. Intenta decir algo como 'Vendí por el valor de 10000'.")
        return None

    accion_detectada = resultado.group(1)
    operacion = identificar_operacion(accion_detectada)
    if operacion is None:
        print(f"❌ No se pudo reconocer la operación: {accion_detectada}")
        return None

    valor = limpiar_valor(resultado.group(2))
    fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    return {
        "fecha_hora": fecha_hora,
        "operacion": operacion,
        "valor": valor
    }

# 🚀 Ejecutar por comando manual con verificación de transacción
while True:
    comando = input("⌨️ Escribe 'e' para registrar una operación o 'salir' para terminar: ").strip().lower()

    if comando == "e":
        texto = escuchar_comando()
        if texto:
            datos = interpretar_comando(texto)
            if datos:
                if confirmar_transaccion(datos['operacion'], datos['valor']):
                    guardar_transaccion(datos['operacion'], datos['valor'])
                else:
                    print("❌ Transacción cancelada por el usuario.")
    elif comando == "salir":
        print("👋 Programa finalizado. ¡Hasta pronto!")
        break
    else:
        print("⚠️ Comando no reconocido. Escribe 'escuchar' o 'salir'.")
